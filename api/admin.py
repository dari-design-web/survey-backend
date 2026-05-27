"""後台管理 API：問卷/題目/填答/兌換券/帳號 CRUD + Excel 匯出"""
import io
import json
import time
from datetime import datetime, timezone
from flask import Blueprint, request, jsonify, g, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill
from lib import store
from lib.auth import require_login, require_role, audit, hash_password
from lib.coupon import void_coupon
from lib.stats import dashboard_stats

bp = Blueprint("api_admin", __name__, url_prefix="/api/admin")


# ============ Dashboard ============
@bp.route("/dashboard", methods=["GET"])
@require_role("super", "admin")
def dashboard():
    return jsonify(dashboard_stats())


# ============ 問卷 CRUD ============
@bp.route("/surveys", methods=["GET"])
@require_role("super", "admin")
def list_surveys():
    rows = []
    for s in store.all_items("surveys"):
        rows.append({**s,
            "question_count": sum(1 for q in store.all_items("survey_questions") if q.get("survey_id") == s["id"]),
            "response_count": sum(1 for r in store.all_items("survey_responses") if r.get("survey_id") == s["id"]),
        })
    rows.sort(key=lambda x: x["id"], reverse=True)
    return jsonify(surveys=rows)


@bp.route("/surveys/<int:sid>", methods=["GET"])
@require_role("super", "admin")
def get_survey(sid):
    s = store.get("surveys", sid)
    if not s:
        return jsonify(error="not_found"), 404
    questions = [q for q in store.all_items("survey_questions") if q.get("survey_id") == sid]
    questions.sort(key=lambda q: (q.get("sort_order", 0), q["id"]))
    return jsonify(survey=s, questions=questions)


def _survey_payload(body):
    return {
        "title": body.get("title") or "未命名問卷",
        "description": body.get("description"),
        "location_name": body.get("location_name"),
        "start_date": body.get("start_date"),
        "end_date": body.get("end_date"),
        "is_active": 1 if body.get("is_active") else 0,
        "generate_coupon": 1 if body.get("generate_coupon") else 0,
        "coupon_valid_days": int(body.get("coupon_valid_days") or 30),
        "redemption_interval_days": int(body.get("redemption_interval_days") or 30),
        "redemption_location": body.get("redemption_location"),
        "coupon_title": body.get("coupon_title") or "贈品兌換券",
        "coupon_subtitle": body.get("coupon_subtitle") or "感謝您完成滿意度調查",
        "thank_you_message": body.get("thank_you_message"),
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }


@bp.route("/surveys", methods=["POST"])
@require_role("super", "admin")
def create_survey():
    body = request.get_json() or {}
    payload = _survey_payload(body)
    payload["created_at"] = datetime.now(timezone.utc).isoformat()
    s = store.insert("surveys", payload)
    audit("create_survey", "survey", s["id"], body)
    return jsonify(ok=True, id=s["id"])


@bp.route("/surveys/<int:sid>", methods=["PUT"])
@require_role("super", "admin")
def update_survey(sid):
    body = request.get_json() or {}
    if not store.get("surveys", sid):
        return jsonify(error="not_found"), 404
    store.update("surveys", sid, _survey_payload(body))
    audit("update_survey", "survey", sid, body)
    return jsonify(ok=True)


@bp.route("/surveys/<int:sid>/duplicate", methods=["POST"])
@require_role("super", "admin")
def duplicate_survey(sid):
    src = store.get("surveys", sid)
    if not src:
        return jsonify(error="not_found"), 404
    payload = {k: v for k, v in src.items() if k != "id"}
    payload["title"] = (payload.get("title") or "") + " (複製)"
    payload["is_active"] = 0
    payload["created_at"] = datetime.now(timezone.utc).isoformat()
    payload["updated_at"] = datetime.now(timezone.utc).isoformat()
    new_s = store.insert("surveys", payload)
    # 複製題目
    for q in store.all_items("survey_questions"):
        if q.get("survey_id") != sid:
            continue
        nq = {k: v for k, v in q.items() if k != "id"}
        nq["survey_id"] = new_s["id"]
        nq["created_at"] = datetime.now(timezone.utc).isoformat()
        store.insert("survey_questions", nq)
    audit("duplicate_survey", "survey", new_s["id"])
    return jsonify(ok=True, id=new_s["id"])


@bp.route("/surveys/<int:sid>/toggle-active", methods=["POST"])
@require_role("super", "admin")
def toggle_active(sid):
    s = store.get("surveys", sid)
    if not s:
        return jsonify(error="not_found"), 404
    nxt = 0 if s.get("is_active") else 1
    store.update("surveys", sid, {"is_active": nxt})
    audit("toggle_survey_active", "survey", sid, {"is_active": nxt})
    return jsonify(ok=True, is_active=nxt)


# ============ 題目 CRUD ============
def _question_payload(body, survey_id=None):
    p = {
        "question_text": body.get("question_text") or "未命名題目",
        "question_type": body.get("question_type") or "text",
        "is_required": 1 if body.get("is_required") else 0,
        "options": body.get("options"),
        "sort_order": int(body.get("sort_order") or 0),
        "is_active": 0 if body.get("is_active") is False else 1,
        "include_in_statistics": 0 if body.get("include_in_statistics") is False else 1,
        "is_satisfaction_score": 1 if body.get("is_satisfaction_score") else 0,
        "is_overall_score": 1 if body.get("is_overall_score") else 0,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }
    if survey_id is not None:
        p["survey_id"] = survey_id
    return p


@bp.route("/surveys/<int:sid>/questions", methods=["POST"])
@require_role("super", "admin")
def create_question(sid):
    body = request.get_json() or {}
    p = _question_payload(body, sid)
    p["created_at"] = datetime.now(timezone.utc).isoformat()
    q = store.insert("survey_questions", p)
    audit("create_question", "question", q["id"])
    return jsonify(ok=True, id=q["id"])


@bp.route("/questions/<int:qid>", methods=["PUT"])
@require_role("super", "admin")
def update_question(qid):
    body = request.get_json() or {}
    if not store.get("survey_questions", qid):
        return jsonify(error="not_found"), 404
    store.update("survey_questions", qid, _question_payload(body))
    audit("update_question", "question", qid)
    return jsonify(ok=True)


@bp.route("/questions/<int:qid>", methods=["DELETE"])
@require_role("super", "admin")
def delete_question(qid):
    store.delete("survey_questions", qid)
    audit("delete_question", "question", qid)
    return jsonify(ok=True)


# ============ 填答資料 ============
@bp.route("/responses", methods=["GET"])
@require_role("super", "admin")
def list_responses():
    survey_id = request.args.get("survey_id", type=int)
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    min_score = request.args.get("min_score", type=float)
    max_score = request.args.get("max_score", type=float)

    responses = store.all_items("survey_responses")
    surveys = {s["id"]: s for s in store.all_items("surveys")}
    users = {u["id"]: u for u in store.all_items("users")}
    questions = {q["id"]: q for q in store.all_items("survey_questions")}
    answers_by_resp = {}
    for a in store.all_items("survey_answers"):
        answers_by_resp.setdefault(a["response_id"], []).append(a)
    coupons_by_resp = {c["response_id"]: c for c in store.all_items("coupons")}

    rows = []
    for r in responses:
        if survey_id and r.get("survey_id") != survey_id:
            continue
        date = (r.get("submitted_at") or "")[:10]
        if date_from and date < date_from:
            continue
        if date_to and date > date_to:
            continue
        overall = None
        for a in answers_by_resp.get(r["id"], []):
            q = questions.get(a.get("question_id"))
            if q and q.get("is_overall_score") and a.get("answer_value") is not None:
                overall = a["answer_value"]
                break
        if min_score is not None and (overall is None or overall < min_score):
            continue
        if max_score is not None and (overall is None or overall > max_score):
            continue
        s = surveys.get(r["survey_id"], {})
        u = users.get(r.get("user_id"), {})
        c = coupons_by_resp.get(r["id"], {})
        rows.append({
            "id": r["id"], "submitted_at": r.get("submitted_at"),
            "survey_id": r["survey_id"], "survey_title": s.get("title"),
            "location_name": s.get("location_name"),
            "user_id": r.get("user_id"),
            "user_name": u.get("name"), "phone": u.get("phone"), "email": u.get("email"),
            "overall_score": overall,
            "coupon_code": c.get("coupon_code"), "coupon_status": c.get("status"),
        })
    rows.sort(key=lambda x: x["id"], reverse=True)
    return jsonify(rows=rows[:500])


@bp.route("/responses/<int:rid>", methods=["GET"])
@require_role("super", "admin")
def get_response(rid):
    r = store.get("survey_responses", rid)
    if not r:
        return jsonify(error="not_found"), 404
    s = store.get("surveys", r["survey_id"]) or {}
    u = store.get("users", r.get("user_id")) or {}
    answers = []
    for a in store.all_items("survey_answers"):
        if a.get("response_id") != rid:
            continue
        q = store.get("survey_questions", a.get("question_id")) or {}
        answers.append({**a,
                        "question_text": q.get("question_text"),
                        "question_type": q.get("question_type"),
                        "sort_order": q.get("sort_order", 0)})
    answers.sort(key=lambda x: x.get("sort_order", 0))
    coupon = store.find_one("coupons", lambda c: c.get("response_id") == rid)
    return jsonify(response={**r,
                              "survey_title": s.get("title"),
                              "location_name": s.get("location_name"),
                              "user_name": u.get("name"),
                              "phone": u.get("phone"), "email": u.get("email")},
                   answers=answers, coupon=coupon)


# ============ 兌換券管理 ============
@bp.route("/coupons", methods=["GET"])
@require_role("super", "admin")
def list_coupons():
    status = request.args.get("status", "")
    search = request.args.get("search", "")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    surveys = {s["id"]: s for s in store.all_items("surveys")}
    users = {u["id"]: u for u in store.all_items("users")}
    admins = {a["id"]: a for a in store.all_items("admin_users")}

    rows = []
    for c in store.all_items("coupons"):
        if status and c.get("status") != status:
            continue
        if search:
            t = search.lower()
            u = users.get(c.get("user_id"), {})
            if (t not in c.get("coupon_code", "").lower()
                and t not in (u.get("phone") or "").lower()
                and t not in (u.get("email") or "").lower()):
                continue
        d = (c.get("issued_at") or "")[:10]
        if date_from and d < date_from:
            continue
        if date_to and d > date_to:
            continue
        s = surveys.get(c.get("survey_id"), {})
        u = users.get(c.get("user_id"), {})
        a = admins.get(c.get("redeemed_by"), {})
        rows.append({**c,
                     "survey_title": s.get("title"),
                     "user_name": u.get("name"), "phone": u.get("phone"), "email": u.get("email"),
                     "redeemer_name": a.get("name")})
    rows.sort(key=lambda x: x["id"], reverse=True)
    return jsonify(rows=rows[:500])


@bp.route("/coupons/<int:cid>/void", methods=["POST"])
@require_role("super", "admin")
def void(cid):
    note = (request.get_json() or {}).get("note") or ""
    r = void_coupon(cid, g.admin_user["id"], note)
    audit("void_coupon", "coupon", cid)
    return jsonify(r)


# ============ 管理者帳號 ============
@bp.route("/admin-users", methods=["GET"])
@require_role("super")
def list_admin_users():
    users = []
    for u in store.all_items("admin_users"):
        users.append({k: v for k, v in u.items() if k != "password_hash"})
    return jsonify(users=users)


@bp.route("/admin-users", methods=["POST"])
@require_role("super")
def create_admin_user():
    body = request.get_json() or {}
    name = body.get("name"); email = body.get("email")
    password = body.get("password"); role = body.get("role")
    if not all([name, email, password, role]):
        return jsonify(error="必填欄位不完整"), 400
    if store.find_one("admin_users", lambda x: x.get("email") == email):
        return jsonify(error="此 Email 已存在"), 400
    u = store.insert("admin_users", {
        "name": name, "email": email,
        "password_hash": hash_password(password),
        "role": role, "is_active": 1,
        "last_login_at": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    audit("create_admin_user", "admin_user", u["id"])
    return jsonify(ok=True, id=u["id"])


@bp.route("/admin-users/<int:uid>", methods=["PUT"])
@require_role("super")
def update_admin_user(uid):
    body = request.get_json() or {}
    if not store.get("admin_users", uid):
        return jsonify(error="not_found"), 404
    patch = {"name": body.get("name"), "role": body.get("role"),
             "is_active": 1 if body.get("is_active") else 0}
    if body.get("password"):
        patch["password_hash"] = hash_password(body["password"])
    store.update("admin_users", uid, patch)
    audit("update_admin_user", "admin_user", uid)
    return jsonify(ok=True)


@bp.route("/admin-users/<int:uid>", methods=["DELETE"])
@require_role("super")
def deactivate_admin_user(uid):
    if uid == g.admin_user["id"]:
        return jsonify(error="不能刪除自己"), 400
    store.update("admin_users", uid, {"is_active": 0})
    audit("deactivate_admin_user", "admin_user", uid)
    return jsonify(ok=True)


# ============ Excel 匯出 ============
def _xlsx(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp.route("/export/responses", methods=["GET"])
@require_role("super", "admin")
def export_responses():
    survey_id = request.args.get("survey_id", type=int)
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    surveys = {s["id"]: s for s in store.all_items("surveys")}
    users = {u["id"]: u for u in store.all_items("users")}
    questions_all = store.all_items("survey_questions")
    answers_all = store.all_items("survey_answers")
    coupons_by_resp = {c["response_id"]: c for c in store.all_items("coupons")}

    responses = []
    for r in store.all_items("survey_responses"):
        if survey_id and r.get("survey_id") != survey_id:
            continue
        d = (r.get("submitted_at") or "")[:10]
        if date_from and d < date_from:
            continue
        if date_to and d > date_to:
            continue
        responses.append(r)
    responses.sort(key=lambda x: x["id"], reverse=True)

    if survey_id:
        qs = [q for q in questions_all if q.get("survey_id") == survey_id]
    else:
        qs = questions_all[:]
    qs.sort(key=lambda q: (q.get("survey_id", 0), q.get("sort_order", 0), q["id"]))

    ans_map = {}
    for a in answers_all:
        ans_map.setdefault(a["response_id"], {})[a["question_id"]] = a

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "問卷填答"
    headers = ["填寫日期", "填寫時間", "問卷名稱", "場域名稱", "使用者識別", "姓名"] + \
              [q["question_text"] for q in qs] + ["兌換券編號", "兌換券狀態"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E7EFFA")
    widths = [12, 10, 25, 20, 22, 12] + [22] * len(qs) + [22, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for r in responses:
        sub = r.get("submitted_at") or ""
        s = surveys.get(r["survey_id"], {})
        u = users.get(r.get("user_id"), {})
        c = coupons_by_resp.get(r["id"], {})
        row = [sub[:10], sub[11:19], s.get("title", ""), s.get("location_name") or "",
               u.get("phone") or u.get("email") or "", u.get("name") or ""]
        for q in qs:
            a = ans_map.get(r["id"], {}).get(q["id"])
            if not a:
                row.append("")
            elif a.get("answer_value") is not None:
                row.append(a["answer_value"])
            else:
                txt = a.get("answer_text") or ""
                try:
                    parsed = json.loads(txt)
                    if isinstance(parsed, list):
                        txt = "、".join(parsed)
                except Exception:
                    pass
                row.append(txt)
        row.extend([c.get("coupon_code", ""), c.get("status", "")])
        ws.append(row)

    audit("export_responses")
    return _xlsx(wb, f"responses_{int(time.time())}.xlsx")


@bp.route("/export/coupons", methods=["GET"])
@require_role("super", "admin")
def export_coupons():
    status = request.args.get("status", "")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    surveys = {s["id"]: s for s in store.all_items("surveys")}
    users = {u["id"]: u for u in store.all_items("users")}
    admins = {a["id"]: a for a in store.all_items("admin_users")}

    # 最近一次兌換日
    last_redeemed = {}
    for c in store.all_items("coupons"):
        if c.get("status") == "redeemed" and c.get("redeemed_at"):
            uid = c["user_id"]
            cur = last_redeemed.get(uid)
            if not cur or c["redeemed_at"] > cur:
                last_redeemed[uid] = c["redeemed_at"]

    rows = []
    for c in store.all_items("coupons"):
        if status and c.get("status") != status:
            continue
        d = (c.get("issued_at") or "")[:10]
        if date_from and d < date_from:
            continue
        if date_to and d > date_to:
            continue
        rows.append(c)
    rows.sort(key=lambda x: x["id"], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "兌換券"
    ws.append(["兌換券編號", "使用者識別", "姓名", "問卷名稱", "產生時間", "有效期限",
               "狀態", "核銷時間", "核銷人員", "最近一次兌換日"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E7EFFA")
    widths = [22, 22, 12, 25, 20, 20, 12, 20, 14, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    for c in rows:
        u = users.get(c.get("user_id"), {})
        s = surveys.get(c.get("survey_id"), {})
        a = admins.get(c.get("redeemed_by"), {})
        ws.append([c["coupon_code"], u.get("phone") or u.get("email") or "",
                   u.get("name") or "", s.get("title") or "",
                   c.get("issued_at"), c.get("expires_at"), c.get("status"),
                   c.get("redeemed_at") or "", a.get("name") or "",
                   last_redeemed.get(c.get("user_id")) or ""])

    audit("export_coupons")
    return _xlsx(wb, f"coupons_{int(time.time())}.xlsx")


@bp.route("/export/summary", methods=["GET"])
@require_role("super", "admin")
def export_summary():
    stats = dashboard_stats()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "統計摘要"
    ws.append(["項目", "數值"])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    for c in ws[1]:
        c.font = Font(bold=True)
    c = stats["cards"]
    for k, v in [
        ("總問卷填寫數", c["totalResponses"]),
        ("今日問卷填寫數", c["todayResponses"]),
        ("本月問卷填寫數", c["monthResponses"]),
        ("平均整體滿意度", c["overallAvg"]),
        ("產生兌換券數", c["totalCoupons"]),
        ("未兌換", c["issuedCoupons"]),
        ("已兌換", c["redeemedCoupons"]),
        ("逾期", c["expiredCoupons"]),
        ("作廢", c["voidedCoupons"]),
        ("兌換率 (%)", c["redemptionRate"]),
        ("本月不滿意回饋數", c["monthNegative"]),
    ]:
        ws.append([k, v])

    ws2 = wb.create_sheet("各滿意度平均")
    ws2.append(["評分項目", "平均", "樣本數"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 30
    for r in stats["satisfactionAvg"]:
        ws2.append([r["label"], r["avg"], r["n"]])

    ws3 = wb.create_sheet("關鍵字統計")
    ws3.append(["類別", "提及次數"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    ws3.column_dimensions["A"].width = 20
    for r in stats["keywords"]:
        ws3.append([r["label"], r["c"]])

    audit("export_summary")
    return _xlsx(wb, f"summary_{int(time.time())}.xlsx")
